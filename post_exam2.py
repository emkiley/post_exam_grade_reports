#!/usr/bin/python
#
# Computes midterm grades in best, worst, and current scenarios, and sends students e-mailed reports
#
# Written by Erin Kiley <emkiley@mcla.edu>, latest revision May 2, 2017

sendmails=1 # 0 to not send e-mails (just print a table for entering midterm grades); 1 to send each student an e-mailed report as well

import openpyxl # for reading Excel files
from openpyxl.utils import get_column_letter, column_index_from_string

if sendmails: # only bother importing MIME/smtp libraries if I want to send e-mails
	import smtplib
	from email.MIMEMultipart import MIMEMultipart
	from email.MIMEText import MIMEText

# Open the grade sheet downloaded from Canvas
wb = openpyxl.load_workbook('10_Apr_14_00_Grades-MATH23201.xlsx')
sheet = wb.active

# Define a function for getting an assignment grade
def getval(i,col): # i is student row number, col is assignment column number
	if type(col)==str:
		val = sheet.cell(row=i,column=column_index_from_string(col)).value
	else:
		val = sheet.cell(row=i,column=col).value
	if val:
		return 1.0*val
	else: # replaces unsubmitted or unmarked with zeros
		return 0.0

# Define a function for computing the grade on an assignment as a percentage of the total
def getavg(i,col): # i is student row number, col is assignment column number
	return getval(i,col)/getval(3,col)

# Define a function for generating e-mail message
def prepare_mail(msgs,anumber,name,exam1,exam2,hwavg,quizavg,excr,summaryavg,discavg,curravg,currletter,worstavg,worstletter,bestavg,bestletter):

    message = "Dear %s,\n\nWe have now completed Exam 2, and I would like you to please check the following averages of your previous graded work against your own records:\n\n\tExam 1 Score: %5.2f%%\n\tExam 2 Score: %5.2f%%\n\tHomework Average: %5.2f%%\n\tQuiz Average: %5.2f%% (with %d out of 3 possible Extra Credit Opportunities completed so far)\n\tWriting Assignments: %5.2f%%\n\tBook Discussions: %5.2f%%\n\nThese values were computed directly using your assignment grades, and not extracted from the averages displayed on Canvas to the right of the assignment grades. If you have found any discrepancies between your own records and the averages listed here, please contact me immediately.\n\nPlease consider the following possible scenarios:\n\n1) In the coming weeks...\n\t* You earn a grade of 0%% on all future homework assignments;\n\t* You earn a grade of 0%% on all future quizzes, and complete no additional Extra Credit Opportunities;\n\t* You earn a grade of 0%% on all future project assignments;\n\t* You earn a grade of 0%% on Exam 3...\n...Then your final course average will be %5.2f, which corresponds to a letter grade of %s.\n\n2) In the coming weeks...\n\t* You earn the same grade as your current homework average on all future homework assignments;\n\t* You earn the same grade as your current quiz average on all future quizzes, and complete no additional Extra Credit Opportunities;\n\t* You earn the same grade as your current book project average on all future project assignments, including the final essay;\n\t* You earn on Exam 3 the average of your grades for Exams 1 and 2...\n...Then your final course average, computed according to the weighting scheme in the syllabus, will be %5.2f, which corresponds to a letter grade of %s.\n\n3) In the coming weeks...\n\t* You earn a grade of 100%% on all future homework assignments;\n\t* You earn a grade of 100%% on all future quizzes, and complete all 3 Extra Credit Opportunities;\n\t* You earn a grade of 100%% on all future project assignments;\n\t* You earn a grade of 100%% on Exam 3...\n...Then your final course average will be %5.2f, which corresponds to a letter grade of %s.\n\nThe grades in each scenario above were computed under the assumption that there will be 21 total homework assignments, 10 total quizzes, and 8 total weeks of the book project. Please realize that anything between (and including) scenario (1) and scenario (3) is possible for you as you go forward in the course.\n\nIf you need them, some resources for improving your grades include:\n\t* The Tuesday night study sessions, 5-6 p.m. in Bowman 205;\n\t* The CSSE office, which is located in Daniels Hall;\n\t* Study groups with your peers;\n\t* Attending office hours if you have specific questions (please see the syllabus for my office location, and days and times of my office hours; if the times there are inconvenient for you, then please send me an e-mail).\n\nIf you have any questions, please do let me know.\n\nBest wishes,\nDr. Kiley" % (name,exam1,exam2,hwavg,quizavg,excr,summaryavg,discavg,worstavg,worstletter,curravg,currletter,bestavg,bestletter)

    msg = MIMEMultipart()
    msg['From'] = "Dr. Kiley <emkiley@mcla.edu>"
    msg['To'] = anumber+"@mcla.edu"
    msg['Cc'] = "Dr. Kiley <emkiley@mcla.edu>"
    msg['Subject'] = "[MATH 232 / Statistics]: Exam 2 Grades and 'Current' Course Average"
    msg.attach(MIMEText(message))

    msgs.append(msg)
    anumbers.append(anumber)

# Sending e-mails--logs in again if server boots you out mid-batch
def try_send(mailServer,username,password,recipient,message,anumber):
	try:
	        mailServer.sendmail(username, recipient, message)
	except Exception as e:
		print("Sending to "+anumbers[i]+" failed")
		print(e)
		if smtplib.SMTPConnectError:
			print("Logging in again...")
		        mailServer = smtplib.SMTP('smtp.outlook.com', 587)
		        mailServer.ehlo()
		        mailServer.starttls()
		        mailServer.ehlo()
		        mailServer.login(username, password)
			try_send(mailServer,username,password,recipient,message,anumber)
	print("Sending to "+anumber+" succeeded")

# Define a function for sending the e-mails
def send_mails(anumbers,msgs):
    import getpass # for soliciting e-mail password
    import sys # for setting recursion limit
    sys.setrecursionlimit(50) # if sending fails due to SMTPConnectError, we want to login again and re-send, but don't want to get stuck in an infinite loop if the server just keeps failing to connect for some reason
    
    print("MCLA e-mail address [example: a30098765@mcla.edu ]:")
    username = raw_input()
    password = getpass.getpass()

    mailServer = smtplib.SMTP('smtp.outlook.com', 587)
    mailServer.ehlo()
    mailServer.starttls()
    mailServer.ehlo()
    mailServer.login(username, password)
    for i in range(0,len(anumbers)):
#	recipient = [anumbers[i]+"@mcla.edu",'emkiley@mcla.edu']
	recipient = ["emkiley@gmail.com"]
        try_send(mailServer,username,password,recipient,msgs[i].as_string(),anumbers[i])
    mailServer.close()

# Define a function for computing course average
def courseavg(hwavg,qzavg,projavg,examavg):
	return 20.0*hwavg + 15.0*qzavg + 48.0*examavg + 17.0*projavg

# Define a function for computing letter grade given numerical average
def lettergrade(avg):
	# Assign letter grade
	if round(avg) < 60:
		return 'F'
	elif round(avg) < 62:
		return 'D-'
	elif round(avg) < 65:
		return 'D'
	elif round(avg) < 67:
		return 'D+'
	elif round(avg) < 69:
		return 'C-'
	elif round(avg) < 76:
		return 'C'
	elif round(avg) < 78:
		return 'C+'
	elif round(avg) < 80:
		return 'B-'
	elif round(avg) < 87:
		return 'B'
	elif round(avg) < 89:
		return 'B+'
	elif round(avg) < 91:
		return 'A-'
	else:
		return 'A'

# Total number of assignments I expect all semester
num_hws = 21
num_quizzes = 10
num_bookweeks = 8

anumbers=[] #initialize empty list of a-numbers
msgs=[] # initialize empty list of messages

# For each student
for i in range(5,48):

	name = sheet.cell(row=i,column=1).value # student name
	anumber = sheet.cell(row=i,column=3).value # student A-number

	# Homework list
	current_hws = [] # list of homework grades
	for j in range(column_index_from_string('G'),column_index_from_string('M')+1):
		current_hws.append(getavg(i,j)) # put grade (out of 1) into list
	
	for j in range(column_index_from_string('O'),column_index_from_string('W')+1):
		current_hws.append(getavg(i,j)) # put grade (out of 1) into list

	# Quiz list
	current_qzs = [] # list of quiz grades
	for j in range(column_index_from_string('AR'),column_index_from_string('AW')+1):
		current_qzs.append(getavg(i,j)) # put grade (out of 1) into list
		current_qzs.sort() # sort in place

	# Extra Credit list
	current_excr = [] # list of extra credit grades
	for j in range(column_index_from_string('AN'),column_index_from_string('AQ')+1):
		current_excr.append(getavg(i,j)) # put grade (out of 1) into list
	num_excr=sum(k>0 for k in current_excr)
	current_excr.sort() # sort in place
	if num_excr>3: # cap the number of extra credit opportunities at 3
		current_excr=current_excr[-3:]
	
	# Project list
	current_projsummaries = []
	for j in range(column_index_from_string('X'),column_index_from_string('AG')+1):
		current_projsummaries.append(getavg(i,j)) # put a grade (out of 1) into list

	current_discussion = []
	for j in range(column_index_from_string('AX'),column_index_from_string('BC')+1):
		current_discussion.append(getavg(i,j)) # put a grade (out of 1) into list
		
	# Exam Average (2 exams so far)
	exam1 = getavg(i,column_index_from_string('BF')) # exam 1
	exam2 = getavg(i,column_index_from_string('BG')) # exam 2
	
	examavg = 0.5*(exam1+exam2) # exam average

	# "Expected" scenario (all future grades are equal to the average of the current grades)
	# Homework
	hwavg=(sum(current_hws)/len(current_hws))
	hws=current_hws

	for j in range(len(current_hws),num_hws): # for all future homeworks
		hws.append(hwavg) # homework grade is average of current homeworks
	hws.sort() # sort homework grades in place
	curr_hwavg=sum(hws[5:])/(len(hws)-5) # drop lowest 5 and average

	# Quizzes
	qzavg=(sum(current_qzs)/len(current_qzs))
	qzs=current_qzs
	for j in range(len(qzs),num_quizzes): # for all future quizzes
		qzs.append(qzavg) # quiz grade is average of current quizzes
	qzs.sort()
	for j in range(0,num_excr): # for the extra credit assignments completed so far
		if 0.75+0.25*current_excr[j] >= qzs[j]: # if quiz grade is lower than extra credit grade
			qzs[j]=0.75+0.25*current_excr[j] # replace quiz grade with extra credit grade
	qzs.sort()
	curr_qzavg = sum(qzs[2:])/(len(qzs)-2) # drop lowest 2 and average
	
	# Book Project
	summaryavg=(sum(current_projsummaries))/(len(current_projsummaries))
	discavg=(sum(current_discussion))/(len(current_discussion))
	summaries=current_projsummaries
	disc=current_discussion
	for j in range(len(current_projsummaries),num_bookweeks): # for all future summaries
		summaries.append(summaryavg) # summary grade is average of current summaries
	for j in range(len(current_discussion),num_bookweeks): # for all future discussions
		disc.append(discavg) # discussion grade is average of current discussions
	curr_projavg=0.75*sum(summaries)/len(summaries)+0.25*sum(disc)/len(disc) # project average
	
	curravg=courseavg(curr_hwavg,curr_qzavg,curr_projavg,examavg)
	currletter=lettergrade(curravg)
	
	# Best scenario (all future grades are 100%)
	# Homework
	hws=current_hws
	for j in range(len(current_hws),num_hws):
		hws.append(1.0)
	hws.sort() # sort homework grades in place
	best_hwavg=sum(hws[5:])/(len(hws)-5) # drop lowest five and average

	# Quizzes
	qzs=current_qzs
	for j in range(len(current_qzs),num_quizzes):
		qzs.append(1.0)
	qzs.sort()
	for j in range(0,3):
		qzs[j]=1.0 # replace quiz grade
	qzs.sort()
	best_qzavg = sum(qzs[2:])/(len(qzs)-2) # drop lowest 2 and average
	
	# Project
	summaries=current_projsummaries
	disc=current_discussion
	for j in range(len(current_projsummaries),num_bookweeks):
		summaries.append(1.0)
	for j in range(len(current_discussion),num_bookweeks):
		disc.append(1.0)
	best_projavg=0.5*sum(summaries)/len(summaries)+0.25*sum(disc)/len(disc)+0.25
	
	# Exams
	best_examavg=(2.0*examavg+1.0)/3.0
	
	bestavg=courseavg(best_hwavg,best_qzavg,best_projavg,best_examavg)
	bestletter=lettergrade(bestavg)
	
	# Worst scenario (all future grades are 0%)
	# Homework
	hws=current_hws
	for j in range(len(current_hws),num_hws):
		hws.append(0.0)
	hws.sort() # sort homework grades in place
	worst_hwavg=sum(hws[5:])/(len(hws)-5) # drop lowest 5 and average

	# Quizzes
	qzs=current_qzs
	for j in range(len(current_qzs),num_quizzes):
		qzs.append(0.0)
	qzs.sort()
	for j in range(0,num_excr): # for the extra credit assignments completed so far
		if 0.75+0.25*current_excr[j] >= qzs[j]: # if quiz grade is lower than extra credit grade
			qzs[j]=0.75+0.25*current_excr[j] # replace quiz grade with extra credit grade
	qzs.sort()
	worst_qzavg = sum(qzs[2:])/(len(qzs)-2) # drop lowest 2 and average

	# Project	
	summaries=current_projsummaries
	disc=current_discussion
	for j in range(len(current_projsummaries),num_bookweeks):
		summaries.append(0.0)
	for j in range(len(current_discussion),num_bookweeks):
		disc.append(0.0)
	worst_projavg=0.5*sum(summaries)/len(summaries)+0.25*sum(disc)/len(disc)+0.25
	
	# Exams
	worst_examavg=2.0*examavg/3.0
	
	worstavg=courseavg(worst_hwavg,worst_qzavg,worst_projavg,worst_examavg)
	worstletter=lettergrade(worstavg)

	# Print me an entry in the table of current midterm grades to enter for each student
	print "\t %s, %s, %5.2f" % (name,currletter,curravg)
	
	# If wanted, prepare an e-mail message to send
	if sendmails:
		prepare_mail(msgs,anumber,name,100*exam1,100*exam2,100*hwavg,100*qzavg,num_excr,100*summaryavg,100*discavg,curravg,currletter,worstavg,worstletter,bestavg,bestletter)

if sendmails: # Send the e-mails
	send_mails(anumbers,msgs)
